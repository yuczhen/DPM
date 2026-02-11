# -*- coding: utf-8 -*-
"""
建立新客戶預測範本 Excel
包含三種客戶類型的範例
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 必填欄位定義
REQUIRED_COLUMNS = [
    'education',
    'month salary',
    'job tenure',
    'post code of permanent address',
    'post code of residential address',
    'residence status',
    'main business',
    'product',
    'loan term',
    'paid installments',
    'debt_to_income_ratio',
    'payment_to_income_ratio',
    'number of overdue before the first month',
    'number of overdue in the first half of the first month',
    'number of overdue in the second half of the first month',
    'number of overdue in the second month',
    'number of overdue in the third month',
    'number of overdue in the fourth month',
    'number of overdue in the fifth month',
    'number of overdue in the sixth month',
    'ID',
    'application date'
]

# 欄位說明
COLUMN_DESCRIPTIONS = {
    'education': '教育程度 (高中(含)以下 / 專科/大學 / 研究所以上)',
    'month salary': '月薪 (單位: 元)',
    'job tenure': '工作年資 (單位: 年，可用小數)',
    'post code of permanent address': '戶籍地郵遞區號 (例: 100, 110, 220)',
    'post code of residential address': '居住地郵遞區號 (例: 100, 110, 220)',
    'residence status': '居住狀況 (own=自有 / rent=租 / family=家人)',
    'main business': '行業別 (例: 製造業, 服務業, 金融業)',
    'product': '貸款產品 (例: Product_A, Standard_Loan)',
    'loan term': '貸款期數 (單位: 月)',
    'paid installments': '已繳期數 (新客戶填 0)',
    'debt_to_income_ratio': 'DTI 債務收入比 (0.35 代表 35%)',
    'payment_to_income_ratio': '還款收入比 (0.15 代表 15%)',
    'number of overdue before the first month': '第0月逾期次數 (新客戶填 0)',
    'number of overdue in the first half of the first month': '第1月上半逾期次數 (新客戶填 0)',
    'number of overdue in the second half of the first month': '第1月下半逾期次數 (新客戶填 0)',
    'number of overdue in the second month': '第2月逾期次數 (新客戶填 0)',
    'number of overdue in the third month': '第3月逾期次數 (新客戶填 0)',
    'number of overdue in the fourth month': '第4月逾期次數 (新客戶填 0)',
    'number of overdue in the fifth month': '第5月逾期次數 (新客戶填 0)',
    'number of overdue in the sixth month': '第6月逾期次數 (新客戶填 0)',
    'ID': '客戶ID (回頭客必填，新客戶留空)',
    'application date': '申請日期 (格式: YYYY-MM-DD，選填)'
}

# 建立三種客戶範例
EXAMPLES = [
    {
        'name': '新客戶 (首次申貸)',
        'education': '專科/大學',
        'month salary': 45000,
        'job tenure': 2.5,
        'post code of permanent address': 100,
        'post code of residential address': 100,
        'residence status': 'rent',
        'main business': '服務業',
        'product': 'Standard_Loan',
        'loan term': 24,
        'paid installments': 0,
        'debt_to_income_ratio': 0.35,
        'payment_to_income_ratio': 0.15,
        'number of overdue before the first month': 0,
        'number of overdue in the first half of the first month': 0,
        'number of overdue in the second half of the first month': 0,
        'number of overdue in the second month': 0,
        'number of overdue in the third month': 0,
        'number of overdue in the fourth month': 0,
        'number of overdue in the fifth month': 0,
        'number of overdue in the sixth month': 0,
        'ID': '',
        'application date': ''
    },
    {
        'name': '回頭客 (無違約記錄)',
        'education': '研究所以上',
        'month salary': 65000,
        'job tenure': 5.0,
        'post code of permanent address': 110,
        'post code of residential address': 110,
        'residence status': 'own',
        'main business': '金融業',
        'product': 'Product_A',
        'loan term': 36,
        'paid installments': 12,
        'debt_to_income_ratio': 0.25,
        'payment_to_income_ratio': 0.12,
        'number of overdue before the first month': 0,
        'number of overdue in the first half of the first month': 0,
        'number of overdue in the second half of the first month': 0,
        'number of overdue in the second month': 0,
        'number of overdue in the third month': 0,
        'number of overdue in the fourth month': 0,
        'number of overdue in the fifth month': 0,
        'number of overdue in the sixth month': 0,
        'ID': 'A123456789',
        'application date': '2024-10-28'
    },
    {
        'name': '回頭客 (有逾期記錄)',
        'education': '高中(含)以下',
        'month salary': 32000,
        'job tenure': 1.0,
        'post code of permanent address': 220,
        'post code of residential address': 220,
        'residence status': 'family',
        'main business': '製造業',
        'product': 'Standard_Loan',
        'loan term': 18,
        'paid installments': 6,
        'debt_to_income_ratio': 0.45,
        'payment_to_income_ratio': 0.20,
        'number of overdue before the first month': 0,
        'number of overdue in the first half of the first month': 1,
        'number of overdue in the second half of the first month': 2,
        'number of overdue in the second month': 1,
        'number of overdue in the third month': 0,
        'number of overdue in the fourth month': 1,
        'number of overdue in the fifth month': 0,
        'number of overdue in the sixth month': 0,
        'ID': 'B987654321',
        'application date': '2024-10-28'
    }
]

# 建立 Excel
wb = Workbook()

# ========== Sheet 1: 欄位說明 ==========
ws1 = wb.active
ws1.title = "欄位說明"

# 標題
ws1['A1'] = 'DPM 違約預測系統 - 欄位說明'
ws1['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws1['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.merge_cells('A1:C1')
ws1.row_dimensions[1].height = 30

# 表頭
ws1['A2'] = '欄位名稱'
ws1['B2'] = '說明'
ws1['C2'] = '必填'
for cell in ['A2', 'B2', 'C2']:
    ws1[cell].font = Font(bold=True, color='FFFFFF')
    ws1[cell].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    ws1[cell].alignment = Alignment(horizontal='center', vertical='center')

# 填入欄位說明
row = 3
for col in REQUIRED_COLUMNS:
    ws1[f'A{row}'] = col
    ws1[f'B{row}'] = COLUMN_DESCRIPTIONS[col]
    if col in ['ID', 'application date']:
        ws1[f'C{row}'] = '選填'
    else:
        ws1[f'C{row}'] = '必填'
    row += 1

# 設定欄寬
ws1.column_dimensions['A'].width = 45
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 10

# ========== Sheet 2: 填寫範例 ==========
ws2 = wb.create_sheet("填寫範例")

# 標題
ws2['A1'] = 'DPM 違約預測系統 - 填寫範例 (三種客戶類型)'
ws2['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws2['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.merge_cells(f'A1:{chr(65+len(REQUIRED_COLUMNS)-1)}1')
ws2.row_dimensions[1].height = 30

# 說明
ws2['A2'] = '範例 1: 新客戶（首次申貸，無歷史記錄）'
ws2['A2'].font = Font(bold=True, color='FF0000')
ws2['A3'] = '範例 2: 回頭客（過去借款正常還款，無違約記錄）'
ws2['A3'].font = Font(bold=True, color='0070C0')
ws2['A4'] = '範例 3: 回頭客（有逾期記錄，高風險）'
ws2['A4'].font = Font(bold=True, color='C00000')

# 表頭 (欄位名稱)
for col_idx, col_name in enumerate(REQUIRED_COLUMNS, start=1):
    cell = ws2.cell(row=5, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 填入範例資料
for row_idx, example in enumerate(EXAMPLES, start=6):
    for col_idx, col_name in enumerate(REQUIRED_COLUMNS, start=1):
        value = example[col_name]
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # 設定顏色標記
        if row_idx == 6:  # 新客戶
            cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        elif row_idx == 7:  # 回頭客無違約
            cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        else:  # 回頭客有逾期
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

# 設定欄寬
for col_idx in range(1, len(REQUIRED_COLUMNS) + 1):
    ws2.column_dimensions[chr(64 + col_idx)].width = 15

# ========== Sheet 3: 空白範本 ==========
ws3 = wb.create_sheet("空白範本")

# 標題
ws3['A1'] = 'DPM 違約預測系統 - 空白範本 (請在此填寫新客戶資料)'
ws3['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws3['A1'].fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.merge_cells(f'A1:{chr(65+len(REQUIRED_COLUMNS)-1)}1')
ws3.row_dimensions[1].height = 30

# 表頭
for col_idx, col_name in enumerate(REQUIRED_COLUMNS, start=1):
    cell = ws3.cell(row=2, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 預留 10 行空白
for row_idx in range(3, 13):
    for col_idx in range(1, len(REQUIRED_COLUMNS) + 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value='')
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 設定欄寬
for col_idx in range(1, len(REQUIRED_COLUMNS) + 1):
    ws3.column_dimensions[chr(64 + col_idx)].width = 15

# 儲存檔案
output_path = r'c:\Users\Shirley\Projects\DPM\Prediction\Prediction data\new_client_template.xlsx'
wb.save(output_path)
print(f"[OK] Template updated: {output_path}")
print(f"   - Sheet 1: Column descriptions ({len(REQUIRED_COLUMNS)} columns)")
print(f"   - Sheet 2: Examples (3 customer types)")
print(f"   - Sheet 3: Blank template (ready to fill)")
